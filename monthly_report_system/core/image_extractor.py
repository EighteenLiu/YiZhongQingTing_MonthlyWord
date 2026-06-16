"""从 xlsx 中提取问题照片，并按锚点行号归类。"""

from __future__ import annotations

import io
from collections import Counter, defaultdict
from pathlib import Path
from typing import DefaultDict, Dict, List, Tuple

from openpyxl import load_workbook
from PIL import Image

from utils.image_utils import save_normalized_image


class ImageExtractError(RuntimeError):
    """图片提取失败。"""


def _cell_text(value: object) -> str:
    """单元格文本标准化。"""

    return str(value or "").strip().replace(" ", "").replace("\n", "")


def _anchor_to_row(image: object) -> int | None:
    """读取 openpyxl 图片锚点所在 Excel 行号。"""

    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return int(marker.row) + 1


def _anchor_to_column(image: object) -> int | None:
    """读取 openpyxl 图片锚点所在 Excel 列号。"""

    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return int(marker.col) + 1


def _image_bytes(image: object) -> bytes:
    """兼容不同 openpyxl 版本读取图片字节。"""

    data = image._data()  # noqa: SLF001 - openpyxl 暂无稳定公开 API
    if isinstance(data, bytes):
        return data
    return data.read()


def _merged_parent_text(worksheet: object, row: int, column: int) -> str:
    """读取合并表头中的父级文本。"""

    for merged_range in worksheet.merged_cells.ranges:
        if row == merged_range.min_row and merged_range.min_col <= column <= merged_range.max_col:
            return _cell_text(worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value)
    return _cell_text(worksheet.cell(row=row, column=column).value)


def _problem_photo_columns(worksheet: object) -> set[int]:
    """识别“问题照片”分组下的图片列。"""

    columns: set[int] = set()
    max_column = int(getattr(worksheet, "max_column", 0) or 0)
    max_header_row = min(3, int(getattr(worksheet, "max_row", 0) or 0))
    for column in range(1, max_column + 1):
        parent_texts = [_merged_parent_text(worksheet, row, column) for row in range(1, max_header_row + 1)]
        header_texts = [_cell_text(worksheet.cell(row=row, column=column).value) for row in range(1, max_header_row + 1)]
        if any("问题照片" in text for text in parent_texts + header_texts):
            columns.add(column)
    return columns


def extract_images_by_row(
    xlsx_path: Path,
    output_dir: Path,
    max_images_per_row: int = 3,
) -> Tuple[Dict[int, List[str]], List[str]]:
    """提取工作簿第一张表中的问题照片，按图片左上角锚点行号归类。

    只保留“问题照片”分组下的图片列，避免把整改照片、二维码等其他图片写入月报。
    """

    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: List[str] = []
    grouped: DefaultDict[int, List[str]] = defaultdict(list)
    seen_by_row: Counter[int] = Counter()

    try:
        workbook = load_workbook(xlsx_path)
    except Exception as exc:
        raise ImageExtractError(f"无法打开 xlsx 提取图片：{exc}") from exc

    worksheet = workbook.worksheets[0]
    problem_photo_columns = _problem_photo_columns(worksheet)
    if not problem_photo_columns:
        warnings.append("未识别到“问题照片”列，未提取任何图片。")
        workbook.close()
        return {}, warnings

    images = getattr(worksheet, "_images", [])
    for index, image in enumerate(images, start=1):
        row_number = _anchor_to_row(image)
        column_number = _anchor_to_column(image)
        if row_number is None or column_number is None:
            warnings.append(f"第 {index} 张图片没有可识别锚点，无法匹配记录。")
            continue
        if column_number not in problem_photo_columns:
            continue

        seen_by_row[row_number] += 1
        if len(grouped[row_number]) >= max_images_per_row:
            continue

        try:
            raw = _image_bytes(image)
            with Image.open(io.BytesIO(raw)) as img:
                jpg_path = save_normalized_image(img, output_dir / f"row_{row_number}_image_{index}.jpg")
            grouped[row_number].append(str(jpg_path))
        except Exception as exc:
            warnings.append(f"第 {index} 张图片提取失败：{exc}")

    for row_number, count in sorted(seen_by_row.items()):
        if count > max_images_per_row:
            warnings.append(f"第 {row_number} 行识别到 {count} 张图片，已只保留前 {max_images_per_row} 张。")

    workbook.close()
    return dict(grouped), warnings
