"""Excel 表格读取；xls 必须先由 Microsoft Excel 转换为 xlsx。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook


class ExcelReadError(RuntimeError):
    """Excel 读取失败。"""


def _converted_path(xls_path: Path, output_dir: Path) -> Path:
    """生成转换后的 xlsx 文件路径。"""

    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{xls_path.stem}_converted.xlsx"


def _find_excel_exe() -> Path | None:
    """查找常见 Microsoft Excel 安装路径，仅用于报错诊断。"""

    candidates = [
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office15\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office15\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office14\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE",
    ]
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            return path
    return None


def _check_excel_com_registered() -> bool:
    """检查 Excel.Application 是否注册到 COM。"""

    try:
        import winreg

        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CLSID"):
            return True
    except Exception:
        return False


def convert_xls_to_xlsx(xls_path: Path, output_dir: Path) -> Path:
    """只使用 Microsoft Excel COM 将 xls 原生另存为 xlsx。"""

    xlsx_path = _converted_path(xls_path, output_dir)
    if xlsx_path.exists():
        xlsx_path.unlink()

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise ExcelReadError("当前环境未安装 pywin32/pythoncom，无法调用 Microsoft Excel 转换 xls。") from exc

    excel_exe = _find_excel_exe()
    registered = _check_excel_com_registered()
    excel: Optional[object] = None
    workbook: Optional[object] = None
    com_initialized = False
    try:
        pythoncom.CoInitialize()
        com_initialized = True

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Open(
            str(xls_path.resolve()),
            0,
            True,
            None,
            "",
            "",
            True,
        )
        workbook.SaveAs(str(xlsx_path.resolve()), FileFormat=51)
    except Exception as exc:  # noqa: BLE001 - COM 异常类型不稳定
        excel_hint = f"检测到 EXCEL.EXE：{excel_exe}" if excel_exe else "未在常见路径检测到 EXCEL.EXE"
        com_hint = "Excel.Application 已注册" if registered else "Excel.Application 未注册"
        raise ExcelReadError(
            "无法调用 Microsoft Excel 将 xls 转换为 xlsx。"
            "本程序不会使用 WPS 或系统默认打开方式。"
            f"{excel_hint}；{com_hint}；底层错误：{exc}"
        ) from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if com_initialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
        raise ExcelReadError("Excel 转换后的 xlsx 文件未生成或为空。")
    return xlsx_path


def _cell_text(value: Any) -> str:
    """单元格文本标准化。"""

    if value is None:
        return ""
    return str(value).strip()


def _header_score(values: list[Any]) -> int:
    """判断某一行像不像真正表头。"""

    keywords = [
        "3级点位",
        "三级点位",
        "4级点位",
        "四级点位",
        "具体问题",
        "检查时间",
        "图片1",
        "问题照片",
        "街道名称",
        "点位名称",
    ]
    joined = "|".join(_cell_text(v).replace(" ", "") for v in values)
    return sum(1 for keyword in keywords if keyword in joined)


def _is_point_header(values: list[Any]) -> bool:
    """判断是否为包含 3级/4级点位的真实业务表头行。"""

    joined = "|".join(_cell_text(v).replace(" ", "") for v in values)
    return ("3级点位" in joined or "三级点位" in joined) and ("4级点位" in joined or "四级点位" in joined)


def _unique_columns(columns: list[str]) -> list[str]:
    """生成 pandas 友好的唯一列名。"""

    seen: dict[str, int] = {}
    result: list[str] = []
    for index, column in enumerate(columns, start=1):
        name = column or f"未命名列{index}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}.{count}")
    return result


def _read_xlsx_with_smart_header(path: Path) -> pd.DataFrame:
    """读取 xlsx，自动处理单行表头和两行表头。

    当前检查表是典型两行表头：第一行有“点位”合并单元格，第二行才有
    “1级点位、2级点位、3级点位、4级点位”。这里会优先识别第二行表头，
    并在第二行为空时回退使用第一行表头。
    """

    workbook = load_workbook(path, data_only=True, read_only=False)
    worksheet = workbook.worksheets[0]
    max_col = worksheet.max_column
    preview_rows = min(5, worksheet.max_row)
    row_values = [
        [worksheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        for row in range(1, preview_rows + 1)
    ]

    point_header_candidates = [index + 1 for index, row in enumerate(row_values) if _is_point_header(row)]
    if point_header_candidates:
        header_row_number = point_header_candidates[0]
    else:
        scores = [_header_score(row) for row in row_values]
        best_score = max(scores)
        # 分数相同时取更靠后的行，兼容“第一行合并大类、第二行真实字段”的表头。
        header_row_number = max(index + 1 for index, score in enumerate(scores) if score == best_score) if best_score > 0 else 1

    if header_row_number > 1:
        parent_values = row_values[header_row_number - 2]
        header_values = row_values[header_row_number - 1]
        columns = []
        for parent, child in zip(parent_values, header_values):
            child_text = _cell_text(child)
            parent_text = _cell_text(parent)
            columns.append(child_text or parent_text)
    else:
        columns = [_cell_text(value) for value in row_values[0]]

    data_start_row = header_row_number + 1
    data = []
    for row_number in range(data_start_row, worksheet.max_row + 1):
        values = [worksheet.cell(row=row_number, column=col).value for col in range(1, max_col + 1)]
        if all(value is None for value in values):
            continue
        data.append(values)

    workbook.close()
    df = pd.DataFrame(data, columns=_unique_columns(columns))
    df.attrs["excel_start_row"] = data_start_row
    df.attrs["header_row"] = header_row_number
    return df


def _read_xls_with_smart_header(path: Path) -> pd.DataFrame:
    """读取 xls 数据，复用与 xlsx 相同的智能表头策略。

    xls 的图片仍由 Excel COM 转成 xlsx 后提取；这里专注于保留原始单元格文本，
    避免某些旧表在另存为 xlsx 后把“具体问题”列规整成占位值。
    """

    raw = pd.read_excel(path, sheet_name=0, header=None, dtype=object)
    if raw.empty:
        raise ExcelReadError("Excel 表格为空，请检查上传文件。")

    preview = raw.iloc[: min(5, len(raw))].values.tolist()
    point_header_candidates = [index + 1 for index, row in enumerate(preview) if _is_point_header(row)]
    if point_header_candidates:
        header_row_number = point_header_candidates[0]
    else:
        scores = [_header_score(row) for row in preview]
        best_score = max(scores)
        header_row_number = max(index + 1 for index, score in enumerate(scores) if score == best_score) if best_score > 0 else 1

    header_values = raw.iloc[header_row_number - 1].tolist()
    if header_row_number > 1:
        parent_values = raw.iloc[header_row_number - 2].tolist()
        columns = [_cell_text(child) or _cell_text(parent) for parent, child in zip(parent_values, header_values)]
    else:
        columns = [_cell_text(value) for value in header_values]

    data_start_row = header_row_number + 1
    data = raw.iloc[data_start_row - 1 :].dropna(how="all")
    df = pd.DataFrame(data.values.tolist(), columns=_unique_columns(columns))
    df.attrs["excel_start_row"] = data_start_row
    df.attrs["header_row"] = header_row_number
    return df


def read_excel_table(path: Path, temp_dir: Path) -> tuple[pd.DataFrame, Path]:
    """读取 Excel 第一张工作表，返回 DataFrame 和后续应使用的 xlsx 路径。"""

    suffix = path.suffix.lower()
    warnings: list[str] = []

    data_path = path
    if suffix == ".xlsx":
        actual_path = path
    elif suffix == ".xls":
        actual_path = convert_xls_to_xlsx(path, temp_dir)
        warnings.append(f"xls 已由 Microsoft Excel 转换为 xlsx：{actual_path.name}。图片提取使用该 xlsx 文件。")
        data_path = path
    else:
        raise ExcelReadError("文件格式不支持，请上传 .xls 或 .xlsx 文件。")

    try:
        if suffix == ".xls":
            try:
                df = _read_xls_with_smart_header(data_path)
                warnings.append("xls 数据已从原文件读取，图片提取使用转换后的 xlsx 文件。")
            except ImportError as exc:
                warnings.append(f"当前环境缺少 xlrd，已回退读取转换后的 xlsx 数据：{exc}")
                df = _read_xlsx_with_smart_header(actual_path)
        else:
            df = _read_xlsx_with_smart_header(actual_path)
    except Exception as exc:
        raise ExcelReadError(f"Excel 无法读取：{exc}") from exc

    if df.empty:
        raise ExcelReadError("Excel 表格为空，请检查上传文件。")

    df.attrs["warnings"] = [*warnings, f"识别到第 {df.attrs.get('header_row', 1)} 行为表头。"]
    return df, actual_path
